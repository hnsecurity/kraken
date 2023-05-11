from stats import levenshtein_distance_stats
from openpyxl.styles import Font
import re
import collections
import openpyxl
import argparse
from stats import *
import logging
import traceback

parser = argparse.ArgumentParser(description='Password stats generator')

parser.add_argument('--pwd', dest='filename_password',
                    help='pasword file', required=True)
parser.add_argument('--out', dest='out_file',
                    help='output file', required=True)
parser.add_argument('--group', dest='filename_users',
                    help='group file', required=False)
parser.add_argument('--regex', dest='regex_file',
                    help='regex file', required=False)
parser.add_argument('--dictionary', dest='dictionary_file',
                    help='dictionary file', required=False)

parser.add_argument('--debug', dest='debug',
                    help='enable debug logs',  action='store_true')
parser.add_argument('--show-not-cracked', dest='show_not_cracked',
                    help='enable not cracked password into the XLSX', action='store_true', required=False)
parser.add_argument('--ignore-case', dest='ignore_case',
                    help='ignore case change in Levenshtein distance', action='store_true', required=False)
parser.add_argument('--pwdump', dest='pwdump_file',
                    help='enable hash count', required=False)

parser.add_argument('--check-leaked', dest='check_leaked',
                    help='check if leaked', action='store_true')
parser.add_argument('--check-leaked-google', dest='check_leaked_google',
                    help='check leaked username,password pair in google database', action='store_true')

args = parser.parse_args()

dictionary_list = []
list_user_password = {}
list_user_password_with_history = {}
list_users = None
list_regex = []
list_user_hash = {}


logger = logging.getLogger('logger')

if args.debug:
    logger.setLevel(logging.DEBUG)
    logger.info('Debug logs enabled')

# load user list
try:
    if args.filename_users:
        with open(args.filename_users) as f:
            list_users = f.read().splitlines()
        logger.info(f'Loaded {len(list_users)} users')
        logger.debug('Loaded users: ' + str(list_users))
except:
    logger.warning(
        'Error while reading users file, skipping...')
    list_users = None

# load regex
try:
    if args.regex_file:
        with open(args.regex_file) as f:
            list_regex = f.read().splitlines()
        logger.info(f'Loaded {len(list_regex)} regexs')
        logger.debug('Loaded regexs: ' + str(list_regex))
except:
    logger.warning('Error while reading regex file, skipping...')
    list_regex = []

# load dictionary
try:
    if args.dictionary_file:
        with open(args.dictionary_file) as f:
            for cur_line in f.readlines():
                cur_line = cur_line.rstrip()
                dictionary_list.append(cur_line)
        logger.info(f'Loaded {len(dictionary_list)} words from dictionary')
        logger.debug('Loaded words: ' + str(dictionary_list))
except:
    logger.warning('Error while reading dictionary file, skipping...')
    dictionary_list = []

# load only password in user list
n_passwords = 0

try:
    with open(args.filename_password) as f:
        for cur_line in f:

            try:
                cur_user, cur_password = cur_line.split(':', 1)
                
                # decode the $HEX[] hashcat format
                m = re.search('^\$HEX\[([0-9a-f]+)\]$',cur_password)
                if m:
                    cur_password = bytes.fromhex(m[1]).decode()
            except:
                continue

            # clean user from history
            cur_user_clean = re.sub("_history[0-9]*$", "", cur_user)

            # insert in list only present users
            if not list_users or cur_user_clean in list_users:
                n_passwords += 1
                if "_history" in cur_user:
                    list_user_password_with_history[cur_user] = cur_password.rstrip(
                        '\r\n')
                else:
                    list_user_password[cur_user] = cur_password.rstrip('\r\n')
except:
    logger.critical('Error while reading password file, exiting...')
    logger.debug(traceback.format_exc())
    exit(1)

logger.info(f'Loaded {n_passwords} passwords')

# load only NTHASH in user list
if args.pwdump_file:
    try:
        n_hashes = 0
        with open(args.pwdump_file) as f:
            for cur_line in f:

                try:
                    cur_user, a, b, cur_password, c = cur_line.split(':', 4)
                except:
                    continue

                # clean user from history
                cur_user_clean = re.sub("_history[0-9]*$", "", cur_user)

                # insert in list only present users
                if not list_users or cur_user_clean in list_users:

                    list_user_hash[cur_user] = cur_password.rstrip('\r\n')
                    n_hashes += 1

        logger.info(f'Loaded {n_hashes} hashes')
    except:
        logger.warning('Error while reading pwdump file, skipping...')


# create the excel workbook
wb = openpyxl.Workbook()

ws_stats = wb.create_sheet('Stats')
ws_hash = wb.create_sheet('Hash')
ws_pwd = wb.create_sheet('Most used')

# remove the default sheet
wb.remove(wb['Sheet'])

ws_stats.column_dimensions['B'].width = 20


# statistics
cur_line = 2

if not list_users is None: 
    logger.debug('Starting password analysis...')
    cur_line = password_analysis_stats(ws_stats, cur_line, args, list_users,
                                    list_user_password)
    cur_line += 5

logger.debug('Starting characters analysis...')
cur_line = characters_analysis_stats(ws_stats, cur_line, args, list_users,
                                     list_user_password)

logger.debug('Starting password length analysis...')
cur_line = password_length_stats(ws_stats, cur_line + 5, args, list_users,
                                 list_user_password)

logger.debug('Starting password topology analysis...')
cur_line = password_topology_stats(
    ws_stats, cur_line + 5, args, list_users, list_user_password, list_regex, dictionary_list)

if args.check_leaked:
    logger.info('Starting leaked password analysis, this could take some time...')
    cur_line = password_leaked_stats(
        ws_stats, cur_line + 5, args, list_users, list_user_password, list_regex, dictionary_list)
else:
    logger.debug('Skipping leaked password analysis...')

if len(list_user_password_with_history) != 0:
    logger.debug('Starting levenshtein distance analysis...')
    cur_line = levenshtein_distance_stats(ws_stats, cur_line + 5, args, list_users,
                                        list_user_password, list_user_password_with_history)
else:
    logger.debug('Skipping levenshtein distance analysis...')

# Password occurences
ws_pwd["A1"] = "Password"
ws_pwd["B1"] = "Occurences"

cur_line = 2

c = collections.Counter(list_user_password[elem]
                        for elem in list_user_password)

for i in c.most_common():
    if i[1] > 3:
        ws_pwd["A" + str(cur_line)] = i[0]
        ws_pwd["B" + str(cur_line)] = i[1]
        cur_line += 1

# Hash occurences
ws_hash["A1"] = "Hash"
ws_hash["B1"] = "Occurences"

cur_line = 2

c = collections.Counter(list_user_hash[elem] for elem in list_user_hash)

for i in c.most_common():
    if i[1] > 3:
        ws_hash["A" + str(cur_line)] = i[0]
        ws_hash["B" + str(cur_line)] = i[1]
        cur_line += 1


logger.info('Saving output file')
wb.save(args.out_file)
